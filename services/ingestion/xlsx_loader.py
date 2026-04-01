"""XLSX ingestion utilities for CloudInsight v2."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None  # type: ignore[assignment]


class XLSXLoader:
    """Load Excel workbooks into a standardized ingestion payload."""

    SUPPORTED_SUFFIXES = {".xlsx"}

    def __init__(self, *, preview_rows: int = 50, max_sheets: int | None = None) -> None:
        self.preview_rows = preview_rows
        self.max_sheets = max_sheets

    def load(self, file_path: str | Path) -> dict[str, Any]:
        """Load XLSX content and return a standardized payload."""
        path = Path(file_path).expanduser()
        detected_type = self._detect_file_type(path)

        try:
            self._validate(path, detected_type)
            self._ensure_dependency()

            workbook = load_workbook(filename=path, read_only=True, data_only=True)
            try:
                total_sheets = len(workbook.sheetnames)
                sheet_names = workbook.sheetnames if self.max_sheets is None else workbook.sheetnames[: self.max_sheets]

                sections: list[dict[str, Any]] = []
                tables: list[dict[str, Any]] = []
                overview_parts: list[str] = []
                warnings: list[str] = []

                for sheet_index, sheet_name in enumerate(sheet_names, start=1):
                    worksheet = workbook[sheet_name]
                    row_count = max((worksheet.max_row or 1) - 1, 0)
                    column_count = worksheet.max_column or 0

                    preview_df = pd.read_excel(path, sheet_name=sheet_name, nrows=self.preview_rows, engine="openpyxl")
                    dtypes = {column: str(dtype) for column, dtype in preview_df.dtypes.items()}
                    summary_text = self._build_sheet_summary(sheet_name, row_count, column_count, dtypes)

                    sections.append(
                        {
                            "id": f"sheet_{sheet_index}",
                            "type": "sheet",
                            "sheet_name": sheet_name,
                            "text": summary_text,
                            "row_count": row_count,
                            "column_count": column_count,
                        }
                    )
                    tables.append(
                        {
                            "id": f"table_{sheet_index}",
                            "type": "table",
                            "name": sheet_name,
                            "columns": preview_df.columns.tolist(),
                            "dtypes": dtypes,
                            "preview_rows": preview_df.to_dict(orient="records"),
                        }
                    )
                    overview_parts.append(summary_text)

                if total_sheets > len(sheet_names):
                    warnings.append(
                        f"Processed {len(sheet_names)} of {total_sheets} sheets. Increase max_sheets to ingest the full workbook."
                    )

                metadata = {
                    "file_size_bytes": path.stat().st_size,
                    "sheet_count": total_sheets,
                    "sheets_processed": len(sheet_names),
                    "preview_row_count_per_sheet": self.preview_rows,
                    "truncated": total_sheets > len(sheet_names),
                }
                content = {
                    "text": "\n".join(overview_parts),
                    "sections": sections,
                    "tables": tables,
                }
                return self._success_response(path, detected_type, metadata, content, warnings)
            finally:
                workbook.close()
        except Exception as exc:  # pragma: no cover - exercised through integration
            return self._error_response(path, detected_type, exc)

    def extract(self, file_path: str | Path) -> dict[str, Any]:
        """Alias to align XLSX ingestion with extractor-style interfaces."""
        return self.load(file_path)

    def _build_sheet_summary(
        self,
        sheet_name: str,
        row_count: int,
        column_count: int,
        dtypes: dict[str, str],
    ) -> str:
        dtype_summary = ", ".join(f"{name} ({dtype})" for name, dtype in dtypes.items()) or "No columns detected"
        return f"Sheet '{sheet_name}' contains {row_count} rows and {column_count} columns: {dtype_summary}."

    def _detect_file_type(self, path: Path) -> str:
        return path.suffix.lower().lstrip(".") or "unknown"

    def _validate(self, path: Path, detected_type: str) -> None:
        if not path.exists():
            raise FileNotFoundError(f"XLSX file not found: {path}")
        if not path.is_file():
            raise ValueError(f"Expected a file path, received: {path}")
        if path.suffix.lower() not in self.SUPPORTED_SUFFIXES:
            raise ValueError(f"Unsupported file type for XLSX loader: {detected_type}")

    def _ensure_dependency(self) -> None:
        if load_workbook is None:
            raise ImportError("XLSX ingestion requires 'openpyxl' to be installed.")

    def _success_response(
        self,
        path: Path,
        detected_type: str,
        metadata: dict[str, Any],
        content: dict[str, Any],
        warnings: list[str],
    ) -> dict[str, Any]:
        return {
            "status": "success",
            "file_name": path.name,
            "file_path": str(path.resolve()),
            "file_type": "xlsx",
            "detected_type": detected_type,
            "metadata": metadata,
            "content": content,
            "warnings": warnings,
            "errors": [],
        }

    def _error_response(self, path: Path, detected_type: str, exc: Exception) -> dict[str, Any]:
        return {
            "status": "error",
            "file_name": path.name,
            "file_path": str(path.resolve()),
            "file_type": "xlsx",
            "detected_type": detected_type,
            "metadata": {},
            "content": {"text": "", "sections": [], "tables": []},
            "warnings": [],
            "errors": [str(exc)],
        }


def load_xlsx(file_path: str | Path, **kwargs: Any) -> dict[str, Any]:
    """Convenience wrapper for XLSX ingestion."""
    return XLSXLoader(**kwargs).load(file_path)
